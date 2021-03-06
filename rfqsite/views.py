from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from .models import CCS_RFQ, RFQ, Part_Header, Burden_Rate, Forecast, Material, Part_Costing, Hardware, Output, Roles, ExtendUser, SP_Master, SP_Set, MC_Master, ACT_Set, MC_Set
from django.utils import timezone
from django.core.files.storage import FileSystemStorage
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.password_validation import validate_password
import os, json, re
from datetime import datetime


def to_float(s):
    s = s.strip()
    return float(s) if s else None

@login_required(login_url='/login')
def rfq_table(request):
    ccs_rfq = CCS_RFQ.objects.all()
    ccs_obj = [obj for obj in ccs_rfq]
    rfq = RFQ.objects.all()
    rfq_obj = [obj.ccs_rfq for obj in rfq]
    for obj in ccs_obj:
        if(obj not in rfq_obj):
            newRFQ = RFQ(ccs_rfq=obj, update_date=datetime.now())
            newRFQ.save()
    projects = RFQ.objects.all()
    context = {
        'projects': projects
    }
    return render(request, 'rfqsite/index.html', context)

@login_required(login_url='/login')
def parts(request, tracker_no):
    if(not RFQ.objects.filter(pk=tracker_no)):
        return HttpResponseNotFound("Page Not Found")
    project = RFQ.objects.get(pk=tracker_no)
    act_set = ACT_Set.objects.filter(tracker_no=project).order_by('mc_id')
    if(len(project.file_path.split('/')) == 3 ):
        file_name = project.file_path.split('/')[2]
    else:
        file_name = ''
    parts = [part for part in Part_Header.objects.filter(tracker_no=tracker_no,level=0)]

    context = {
        'act_set': act_set,
        'project': project,
        'parts': parts,
        'file_name': file_name
    }
    return render(request, 'rfqsite/part_table.html', context)

@login_required(login_url='/login')
def rfq_summary(request, tracker_no):
    if(not RFQ.objects.filter(pk=tracker_no)):
        return HttpResponseNotFound("Page Not Found")
    project = RFQ.objects.get(pk=tracker_no)
    total_cost_td = 0
    mtl_cost_td = 0
    spl_process_cost_td = 0
    total_hardware_cost_td = 0
    total_machine_cost_td = 0
    total_nre_cost_td = 0
    hardware_burden_cost_td = 0
    sp_burden_cost_td = 0
    material_burden_cost_td = 0
    all_part = Part_Header.objects.filter(tracker_no=project)
    for item in all_part:
        op = Output.objects.get(sl_no=item)
        pc = Part_Costing.objects.get(sl_no=item)
        total_cost_td += op.total_cost
        mtl_cost_td += op.mtl_cost
        spl_process_cost_td += op.spl_process_cost
        total_hardware_cost_td += op.total_hardware_cost
        total_machine_cost_td += op.total_machine_cost
        total_nre_cost_td += pc.total_nre_cost
        hardware_burden_cost_td += op.hardware_burden_cost
        sp_burden_cost_td += op.sp_burden_cost
        material_burden_cost_td += op.material_burden_cost
    all_act = ACT_Set.objects.filter(tracker_no=project)
    sum = {}
    for act in all_act:
        amc = MC_Set.objects.filter(act_id=act)
        for mc in amc:
            if mc.act_id.mc_id.name in sum:
                sum[mc.act_id.mc_id.name] += mc.mcrftp
            else:
                sum[mc.act_id.mc_id.name] = mc.mcrftp
    mc_set = [MC(x,sum[x]) for x in sum]
    ## structure
    parts = []
    ph = Part_Header.objects.filter(tracker_no=tracker_no,level=0)
    for item in ph:
        sl_no = item.sl_no
        base_level = Part_Header.objects.get(pk=sl_no)
        while(base_level.level > 0):
            base_level = base_level.parent_sl_no
        tree = Part_Tree(base_level,item)
        tree.set_tree()
        # tree.set_open()
        parts.append(tree)
    context = {
        'project': project,
        'total_cost_td': total_cost_td,
        'mtl_cost_td': mtl_cost_td,
        'spl_process_cost_td': spl_process_cost_td,
        'total_hardware_cost_td': total_hardware_cost_td,
        'total_machine_cost_td': total_machine_cost_td,
        'total_nre_cost_td': total_nre_cost_td,
        'material_burden_cost_td': material_burden_cost_td,
        'sp_burden_cost_td': sp_burden_cost_td,
        'hardware_burden_cost_td': hardware_burden_cost_td,
        'mc_set': mc_set,
        'parts': parts
    }
    return render(request, 'rfqsite/rfq_summary.html', context)

@login_required(login_url='/login')
def rfq_summary2(request, tracker_no, sl_no):
    if(not RFQ.objects.filter(pk=tracker_no)):
        return HttpResponseNotFound("Page Not Found")
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    part = Part_Header.objects.get(pk=sl_no)
    project = RFQ.objects.get(pk=tracker_no)
    current_sl = sl_no
    total_cost_td = 0
    mtl_cost_td = 0
    spl_process_cost_td = 0
    total_hardware_cost_td = 0
    total_machine_cost_td = 0
    total_nre_cost_td = 0
    hardware_burden_cost_td = 0
    sp_burden_cost_td = 0
    material_burden_cost_td = 0
    all_part = []
    c = [Part_Header.objects.get(pk=sl_no)]
    while(len(c) > 0):
        for item in c:
            all_part.append(item)
            fo = Part_Header.objects.filter(parent_sl_no=item)
            for item2 in fo:
                c.append(item2)
            c.remove(item)
    for item in all_part:
        op = Output.objects.get(sl_no=item)
        pc = Part_Costing.objects.get(sl_no=item)
        total_cost_td += op.total_cost
        mtl_cost_td += op.mtl_cost
        spl_process_cost_td += op.spl_process_cost
        total_hardware_cost_td += op.total_hardware_cost
        total_machine_cost_td += op.total_machine_cost
        total_nre_cost_td += pc.total_nre_cost
        hardware_burden_cost_td += op.hardware_burden_cost
        sp_burden_cost_td += op.sp_burden_cost
        material_burden_cost_td += op.material_burden_cost
    all_act = ACT_Set.objects.filter(tracker_no=project)
    sum = {}
    for act in all_act:
        amc = MC_Set.objects.filter(act_id=act)
        for mc in amc:
            if(mc.sl_no in all_part):
                if mc.act_id.mc_id.name in sum:
                    sum[mc.act_id.mc_id.name] += mc.mcrftp
                else:
                    sum[mc.act_id.mc_id.name] = mc.mcrftp
    mc_set = [MC(x,sum[x]) for x in sum]
    ## structure
    parts = []
    ph = Part_Header.objects.filter(tracker_no=tracker_no,level=0)
    for item in ph:
        sl_no = item.sl_no
        base_level = Part_Header.objects.get(pk=sl_no)
        while(base_level.level > 0):
            base_level = base_level.parent_sl_no
        tree = Part_Tree(base_level,Part_Header.objects.get(pk=current_sl))
        tree.set_tree()
        if tree.base_level in tree.get_path():
            tree.set_open()
        # print(tree.get_path())
        parts.append(tree)
    context = {
        'project': project,
        'part': part,
        'total_cost_td': total_cost_td,
        'mtl_cost_td': mtl_cost_td,
        'spl_process_cost_td': spl_process_cost_td,
        'total_hardware_cost_td': total_hardware_cost_td,
        'total_machine_cost_td': total_machine_cost_td,
        'total_nre_cost_td': total_nre_cost_td,
        'material_burden_cost_td': material_burden_cost_td,
        'sp_burden_cost_td': sp_burden_cost_td,
        'hardware_burden_cost_td': hardware_burden_cost_td,
        'mc_set': mc_set,
        'parts': parts
    }
    return render(request, 'rfqsite/rfq_summary.html', context)

class MC:
    def __init__(self, name, mcrftp):
        self.name = name
        self.mcrftp = mcrftp

@login_required(login_url='/login')
def add_part(request, tracker_no):
    if(not RFQ.objects.filter(pk=tracker_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    project = RFQ.objects.get(pk=tracker_no)
    context = {
        'project': project
    }
    return render(request, 'rfqsite/add_part.html', context)


@login_required(login_url='/login')
def add_part_multi(request, tracker_no):
    if(not RFQ.objects.filter(pk=tracker_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    project = RFQ.objects.get(pk=tracker_no)
    context = {
        'project': project
    }
    return render(request, 'rfqsite/add_part_multi.html', context)

@login_required(login_url='/login')
def add_child_part_multi(request, sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    part = Part_Header.objects.get(pk=sl_no)
    project = part.tracker_no
    child_level = part.level + 1
    context = {
            'project': project,
            'part': part,
            'child_level': child_level
    }
    return render(request, 'rfqsite/add_child_part_multi.html', context)

@login_required(login_url='/login')
def add_child_part_multi_confirm(request):
    if request.method == 'POST':
        sl_no = request.POST.get('parent-sl-no')
        tracker_no = request.POST.get('tracker-no')
        parent_sl_no = request.POST.get('parent-sl-no')
        part_level = request.POST.get('part-level')
        excel = request.POST.get('excel-data')
        excel = excel.strip()
        trans = re.split('	|\r\n',excel)
        if(len(trans)%10 != 0):
            return redirect('/part_info/'+sl_no+"?message=0&tab=2")
        for i in range(0, len(trans), 10):
            try:
                trans[i] = str(trans[i])
                trans[i+1] = str(trans[i+1])
                trans[i+2] = str(trans[i+2])
                trans[i+3] = str(trans[i+3])
                if(trans[i+3].lower() != ('hardware')):
                    if(trans[i+3].lower() != ('part')):
                        return redirect('/part_info/'+str(sl_no)+"?message=0&tab=2")
                trans[i+4] = int(trans[i+4].replace(',',''))
                trans[i+5] = int(trans[i+5].replace(',',''))
                trans[i+6] = int(trans[i+6].replace(',',''))
                trans[i+7] = int(trans[i+7].replace(',',''))
                trans[i+8] = int(trans[i+8].replace(',',''))
                trans[i+9] = int(trans[i+9].replace(',',''))
            except:
                return redirect('/part_info/'+str(sl_no)+"?message=0&tab=2")
        for i in range(0, len(trans), 10):
            type = trans[i+3].lower().capitalize()
            newPart = Part_Header(tracker_no=RFQ.objects.get(tracker_no=tracker_no),parent_sl_no=Part_Header.objects.get(pk=sl_no),level=part_level,no=trans[i],name=trans[i+1],program=trans[i+2],type=type)
            newForecast = Forecast(sl_no=newPart)
            newForecast.qty_per_unit = trans[i+4]
            newForecast.forecast_year1 = trans[i+5]
            newForecast.forecast_year2 = trans[i+6]
            newForecast.forecast_year3 = trans[i+7]
            newForecast.forecast_year4 = trans[i+8]
            newForecast.forecast_year5 = trans[i+9]
            newMaterial = Material(sl_no=newPart)
            newPart_Costing = Part_Costing(sl_no=newPart)
            newBR = Burden_Rate(sl_no=newPart)
            if(newPart.type == 'Hardware'):
                newBR.material = 0
                newBR.sp = 0
                newBR.subcontract = 0
            else:
                newBR.hardware = 0
            newHW = Hardware(sl_no=newPart)
            newO = Output(sl_no=newPart)
            newPart.save()
            newForecast.save()
            newMaterial.save()
            newPart_Costing.save()
            newBR.save()
            newHW.save()
            newO.save()
        return redirect('/part_info/'+str(sl_no)+"?message=1&tab=2")

#725Z2463-101	FCL_CLEVIS_OIL-TANK-ACCESS-DOOR ASSY.	101	101	101	101	101
#725Z2463-111	FCL_CLEVIS_OIL-TANK-ACCESS-DOOR	101	101	101	101	101
#NAS77C4-014	BUSH	101	101	101	101	101
@login_required(login_url='/login')
def add_part_multi_confirm(request):
    if request.method == 'POST':
        tracker_no = request.POST.get('tracker-no')
        part_level = request.POST.get('part-level')
        excel = request.POST.get('excel-data')
        excel = excel.strip()
        trans = re.split('	|\r\n',excel)
        if(len(trans)%9 != 0):
            return redirect('/part_table/'+request.POST.get('tracker-no')+"?message=0")
        for i in range(0, len(trans), 9):
            try:
                trans[i] = str(trans[i])
                trans[i+1] = str(trans[i+1])
                trans[i+2] = int(trans[i+2].replace(',',''))
                trans[i+3] = int(trans[i+3].replace(',',''))
                trans[i+4] = int(trans[i+4].replace(',',''))
                trans[i+5] = int(trans[i+5].replace(',',''))
                trans[i+6] = int(trans[i+6].replace(',',''))
                trans[i+7] = int(trans[i+7].replace(',',''))
                trans[i+8] = float(trans[i+8].replace(',',''))
            except:
                return redirect('/part_table/'+request.POST.get('tracker-no')+"?message=0")
        for i in range(0, len(trans), 9):
            newPart = Part_Header(tracker_no=RFQ.objects.get(tracker_no=tracker_no),level=part_level,no=trans[i],name=trans[i+1])
            newForecast = Forecast(sl_no=newPart)
            newForecast.qty_per_unit = trans[i+2]
            newForecast.forecast_year1 = trans[i+3]
            newForecast.forecast_year2 = trans[i+4]
            newForecast.forecast_year3 = trans[i+5]
            newForecast.forecast_year4 = trans[i+6]
            newForecast.forecast_year5 = trans[i+7]
            newMaterial = Material(sl_no=newPart)
            newPart_Costing = Part_Costing(sl_no=newPart)
            newPart_Costing.target_price = trans[i+8]
            newBR = Burden_Rate(sl_no=newPart)
            newHW = Hardware(sl_no=newPart)
            newO = Output(sl_no=newPart)
            newPart.save()
            newForecast.save()
            newMaterial.save()
            newPart_Costing.save()
            newBR.save()
            newHW.save()
            newO.save()
        return redirect('/part_table/'+request.POST.get('tracker-no')+"?message=1")

@login_required(login_url='/login')
def add_part_confirm(request):
    if request.method == 'POST':
        tracker_no = request.POST.get('tracker-no')
        part_level = request.POST.get('part-level')
        part_no = request.POST.get('part-no')
        part_name = request.POST.get('part-name')
        program = request.POST.get('program')
        newPart = Part_Header(tracker_no=RFQ.objects.get(tracker_no=tracker_no),level=part_level,no=part_no,name=part_name,program=program)
        newForecast = Forecast(sl_no=newPart)
        newMaterial = Material(sl_no=newPart)
        newPart_Costing = Part_Costing(sl_no=newPart)
        newBR = Burden_Rate(sl_no=newPart)
        newHW = Hardware(sl_no=newPart)
        newO = Output(sl_no=newPart)
        newPart.save()
        newForecast.save()
        newMaterial.save()
        newPart_Costing.save()
        newBR.save()
        newHW.save()
        newO.save()
        return redirect('/part_table/'+tracker_no+"?message=1")

@login_required(login_url='/login')
def edit_rfq(request, tracker_no):
    if(not RFQ.objects.filter(pk=tracker_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    project = RFQ.objects.get(pk=tracker_no)
    if(len(project.file_path.split('/')) == 3 ):
        file_name = project.file_path.split('/')[2]
    else:
        file_name = ''
    remark_time = project.remark_date.strftime("%Y-%m-%dT%H:%M")
    print(remark_time)
    context = {
            'project': project,
            'file_name': file_name,
            'remark_time': remark_time
    }
    return render(request, 'rfqsite/edit_rfq.html', context)

@login_required(login_url='/login')
def edit_rfq_confirm(request):
    if request.method == 'POST' and request.FILES:
        file = request.FILES['file']
        fs = FileSystemStorage()
        tracker_no = request.POST.get('tracker-no')
        project = RFQ.objects.get(pk=tracker_no)
        if(project.file_path != ''):
            if(os.path.exists(os.getcwd().replace('\\','/')+"/media/"+project.file_path)):
                os.remove(os.getcwd().replace('\\','/')+"/media/"+project.file_path)
            project.file_path = ''
        folder = "rfq/"+tracker_no+"/"
        path = folder+file.name
        filename = fs.save(path, file)
        project.file_path = path
        project.usd_thb = request.POST.get('usd-thb')
        project.current_year = request.POST.get('current-year')
        project.hyper_link = request.POST.get('hyper-link')
        project.remark = request.POST.get('remark')
        project.status = request.POST.get('status')
        if(request.POST.get('remark-date') != ""):
            project.remark_date = request.POST.get('remark-date')
        if(request.POST.get('usd-thb') == ''):
            project.usd_thb = 35
        else:
            project.usd_thb = request.POST.get('usd-thb')
        project.save()
        return redirect('/part_table/'+tracker_no+"?message=1")
    elif request.method == 'POST':
        tracker_no = request.POST.get('tracker-no')
        project = RFQ.objects.get(pk=tracker_no)
        # if(request.POST.get('file') == ''):
        #     if(project.file_path != ''):
        #         os.remove(os.getcwd().replace('\\','/')+"/media/"+project.file_path)
        #         project.file_path = ''
        if(request.POST.get('usd-thb') == ''):
            project.usd_thb = 35
        else:
            project.usd_thb = request.POST.get('usd-thb')
        project.hyper_link = request.POST.get('hyper-link')
        project.remark = request.POST.get('remark')
        project.status = request.POST.get('status')
        if(request.POST.get('remark-date') != ""):
            project.remark_date = request.POST.get('remark-date')
        project.current_year = request.POST.get('current-year')
        project.save()
        return redirect('/part_table/'+tracker_no+"?message=1")


@login_required(login_url='/login')
def part_info(request, sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    part = Part_Header.objects.get(pk=sl_no)
    current_year = part.tracker_no.current_year
    base_level = Part_Header.objects.get(pk=sl_no)
    while(base_level.level > 0):
        base_level = base_level.parent_sl_no
    tree = Part_Tree(base_level,part)
    tree.set_tree()
    tree.set_open()
    parts = [tree]
    mc_set = MC_Set.objects.filter(sl_no=part)
    forecast = Forecast.objects.get(sl_no=part)
    material = Material.objects.get(sl_no=part)
    part_costing = Part_Costing.objects.get(sl_no=part)
    burden_rate = Burden_Rate.objects.get(sl_no=part)
    hardware = Hardware.objects.get(sl_no=part)
    sp_set = SP_Set.objects.filter(sl_no=part).order_by('sp_id')
    sum_child = 0
    fchild = Part_Header.objects.filter(parent_sl_no=part)
    child = [x for x in fchild]
    while(len(child) > 0):
        newChild = []
        for item in child:
            op = Output.objects.get(sl_no=item)
            sum_child += op.ccs_ewp
            addChild = [x for x in Part_Header.objects.filter(parent_sl_no=item)]
            for c in addChild:
                 newChild.append(c)
        child = newChild
    if(len(part.file_path.split('/')) == 3 ):
        file_name = part.file_path.split('/')[2]
    else:
        file_name = ''
    data_set = []
    context = {
        'data_set': data_set,
        'sp_set': sp_set,
        'mc_set': mc_set,
        'current_year': current_year,
        'part': part,
        'parts': parts,
        'forecast': forecast,
        'material': material,
        'part_costing': part_costing,
        'burden_rate': burden_rate,
        'hardware': hardware,
        'sum_child': sum_child,
        'file_name': file_name
    }
    return render(request, 'rfqsite/part_info.html', context)

@login_required(login_url='/login')
def edit_part_info(request, sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    part = Part_Header.objects.get(sl_no=sl_no)
    if(len(part.image_path.split('/')) == 3 ):
        image_name = part.image_path.split('/')[2]
    else:
        image_name = ''
    if(len(part.file_path.split('/')) == 3 ):
        file_name = part.file_path.split('/')[2]
    else:
        file_name = ''
    context = {
            'part': part,
            'file_name': file_name,
            'image_name': image_name,
    }
    return render(request, 'rfqsite/edit_part_info.html', context)

@login_required(login_url='/login')
def edit_part_info_confirm(request):
    sl_no = request.POST.get('sl-no')
    if request.method == 'POST':
        editPart = Part_Header.objects.get(sl_no=sl_no)
        print(request.POST)
        # if(request.POST.get('file') == ''):
        #     if(editPart.file_path != ''):
        #         if(os.path.exists(os.getcwd().replace('\\','/')+"/media/"+editPart.file_path)):
        #             os.remove(os.getcwd().replace('\\','/')+"/media/"+editPart.file_path)
        #         editPart.file_path = ''
        # if(request.POST.get('image') == ''):
        #     if(editPart.image_path != ''):
        #         if(os.path.exists(os.getcwd().replace('\\','/')+"/media/"+editPart.image_path)):
        #             os.remove(os.getcwd().replace('\\','/')+"/media/"+editPart.image_path)
        #         editPart.image_path = ''
        editPart.no = request.POST.get('part-no')
        editPart.name = request.POST.get('part-name')
        editPart.program = request.POST.get('program')
        editPart.save()
    if request.FILES:
        files = []
        for key in request.FILES:
            files.append(key)
        if 'file' in files:
            file = request.FILES['file']
            fs = FileSystemStorage()
            if(editPart.file_path != ''):
                os.remove(os.getcwd().replace('\\','/')+"/media/"+editPart.file_path)
                editPart.file_path = ''
            folder = "part/"+sl_no+"/"
            path = folder+file.name
            filename = fs.save(path, file)
            editPart = Part_Header.objects.get(sl_no=sl_no)
            editPart.file_path = path
            editPart.save()
        if 'image' in files:
            file = request.FILES['image']
            fs = FileSystemStorage()
            if(editPart.image_path != ''):
                os.remove(os.getcwd().replace('\\','/')+"/media/"+editPart.image_path)
                editPart.image_path = ''
            folder = "image/"+sl_no+"/"
            path = folder+file.name
            filename = fs.save(path, file)
            editPart = Part_Header.objects.get(sl_no=sl_no)
            editPart.image_path = path
            editPart.save()
    return redirect('/part_info/'+str(sl_no)+"?message=1")

@login_required(login_url='/login')
def edit_forecast(request, sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    forecast = Forecast.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
    part_costing = Part_Costing.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
    part = Part_Header.objects.get(pk=sl_no)
    current_year = part.tracker_no.current_year
    context = {
        'part_costing': part_costing,
        'current_year': current_year,
        'part': part,
        'forecast': forecast
    }
    return render(request, 'rfqsite/edit_forecast.html', context)

@login_required(login_url='/login')
def edit_forecast_confirm(request):
    sl_no = request.POST.get('sl-no')
    if request.method == 'POST':
        editForecast = Forecast.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
        editForecast.qty_per_unit = request.POST.get('qty-per-unit')
        editForecast.forecast_year1 = request.POST.get('forecast-year1')
        editForecast.forecast_year2 = request.POST.get('forecast-year2')
        editForecast.forecast_year3 = request.POST.get('forecast-year3')
        editForecast.forecast_year4 = request.POST.get('forecast-year4')
        editForecast.forecast_year5 = request.POST.get('forecast-year5')
        editForecast.aeqfc = request.POST.get('aeqfc')
        editpc = Part_Costing.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
        editpc.ebq_customer_qty = to_float(request.POST.get('ebq-customer-qty'))
        editpc.ebq_ccs_qty = to_float(request.POST.get('ebq-ccs-qty'))
        editForecast.save()
        editpc.save()
    return redirect('/part_info/'+sl_no+"?message=1&tab=3")


@login_required(login_url='/login')
def edit_hardware(request, sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    hardware = Hardware.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
    part = Part_Header.objects.get(pk=sl_no)
    context = {
        'part': part,
        'hardware': hardware
    }
    return render(request, 'rfqsite/edit_hardware.html', context)

@login_required(login_url='/login')
def edit_hardware_confirm(request):
    sl_no = request.POST.get('sl-no')
    if request.method == 'POST':
        edithw = Hardware.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
        edithw.description = request.POST.get('hardware-description')
        edithw.supplier = request.POST.get('hardware-supplier')
        edithw.price = to_float(request.POST.get('hardware-price'))
        edithw.qty = to_float(request.POST.get('hardware-qty'))
        edithw.save()
        return redirect('/part_info/'+sl_no+"?message=1&tab=5")

@login_required(login_url='/login')
def edit_part_costing(request, sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    part_costing = Part_Costing.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
    part = Part_Header.objects.get(pk=sl_no)
    context = {
        'part': part,
        'part_costing': part_costing
    }
    return render(request, 'rfqsite/edit_part_costing.html', context)

@login_required(login_url='/login')
def edit_part_costing_confirm(request):
    sl_no = request.POST.get('sl-no')
    if request.method == 'POST':
        editpc = Part_Costing.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
        editpc.total_nre_cost = to_float(request.POST.get('total-nre-cost'))
        editpc.target_price = to_float(request.POST.get('target-price'))
        editpc.otspsuc = to_float(request.POST.get('otspsuc'))
        editpc.otrmac = to_float(request.POST.get('otrmac'))
        editpc.ottc = to_float(request.POST.get('ottc'))
        editpc.ccs_quote_assumptions = request.POST.get('ccs-quote-assumptions')
        editpc.base_subcontract = to_float(request.POST.get('base-subcontract'))
        editpc.shipping_cost = to_float(request.POST.get('shipping-cost'))
        editpc.save()
        return redirect('/part_info/'+sl_no+"?message=1&tab=7")

@login_required(login_url='/login')
def edit_dltiw(request, sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    part_costing = Part_Costing.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
    part = Part_Header.objects.get(pk=sl_no)
    context = {
        'part': part,
        'part_costing': part_costing
    }
    return render(request, 'rfqsite/edit_dltiw.html', context)

@login_required(login_url='/login')
def edit_dltiw_confirm(request):
    sl_no = request.POST.get('sl-no')
    if request.method == 'POST':
        editpc = Part_Costing.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
        editpc.dltiw_fai = request.POST.get('dltiw-fai')
        editpc.dltiw_serial_production = request.POST.get('dltiw-serial-production')
        editpc.dltiw_production = request.POST.get('dltiw-production')
        editpc.save()
        return redirect('/part_info/'+sl_no+"?message=1&tab=7")


@login_required(login_url='/login')
def add_child(request, sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    part = Part_Header.objects.get(pk=sl_no)
    project = part.tracker_no
    child_level = part.level + 1
    context = {
            'project': project,
            'part': part,
            'child_level': child_level
    }
    return render(request, 'rfqsite/add_child.html', context)

@login_required(login_url='/login')
def add_child_confirm(request):
    if request.method == 'POST':
        sl_no = request.POST.get('sl-no')
        tracker_no = request.POST.get('tracker-no')
        parent_sl_no = request.POST.get('parent-sl-no')
        part_level = request.POST.get('part-level')
        part_no = request.POST.get('part-no')
        part_name = request.POST.get('part-name')
        program = request.POST.get('program')
        fco = Forecast.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
        type = request.POST.get('type')
        newPart = Part_Header(type=type,tracker_no=RFQ.objects.get(tracker_no=tracker_no),level=part_level,no=part_no,name=part_name,program=program, parent_sl_no=Part_Header.objects.get(pk=sl_no))
        newForecast = Forecast(sl_no=newPart, forecast_year1=fco.forecast_year1, forecast_year2=fco.forecast_year2, forecast_year3=fco.forecast_year3, forecast_year4=fco.forecast_year4, forecast_year5=fco.forecast_year5)
        newMaterial = Material(sl_no=newPart)
        newPart_Costing = Part_Costing(sl_no=newPart)
        newBR = Burden_Rate(sl_no=newPart)
        if(type == 'Hardware'):
            newBR.material = 0
            newBR.sp = 0
            newBR.subcontract = 0
        else:
            newBR.hardware = 0
        newHW = Hardware(sl_no=newPart)
        newO = Output(sl_no=newPart)
        newPart.save()
        newForecast.save()
        newMaterial.save()
        newPart_Costing.save()
        newBR.save()
        newHW.save()
        newO.save()
        return redirect('/part_info/'+sl_no+"?message=1&tab=2")

@login_required(login_url='/login')
def edit_material(request, sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    material = Material.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
    part = Part_Header.objects.get(pk=sl_no)
    context = {
        'part': part,
        'material': material
    }
    return render(request, 'rfqsite/edit_material.html', context)

def clean_material(material):
    material.description = ''
    material.cross_section = ''
    material.type = ''
    material.quantity = 1
    material.rm_density = 0
    material.rm_density_unit = ''
    material.rm_d1 = 0
    material.rm_d1_unit = ''
    material.rm_d2 = 0
    material.rm_d2_unit = ''
    material.rm_t = 0
    material.rm_t_unit = ''
    material.rm_w = 0
    material.rm_w_unit = ''
    material.rm_l = 0
    material.rm_l_unit = ''
    material.rm_total_weight = 0
    material.supplier = ''
    material.base_material_price = 0
    material.shipping_cost = 0
    material.fpwp = 0
    material.armwpp = 0
    material.save()

def to_str(str):
    if str == None:
        return ''
    else:
        return str

@login_required(login_url='/login')
def edit_material_confirm(request):
    sl_no = request.POST.get('sl-no')
    if request.method == 'POST':
        editmat = Material.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
        clean_material(editmat)
        type = request.POST.get('cross-section')
        editmat.description = to_str(request.POST.get('description'))
        if(request.POST.get('check-armwpp') == None):
            editmat.cross_section = to_str(request.POST.get('cross-section'))
            editmat.type = to_str(request.POST.get('type'))
            editmat.quantity = to_float(request.POST.get('quantity'))
            editmat.rm_density = to_float(request.POST.get('density'))
            editmat.rm_density_unit = request.POST.get('density-unit')
            editmat.rm_total_weight = request.POST.get('rm-total-weight')
        if(request.POST.get('check-armwpp') != None):
            editmat.armwpp = to_float(request.POST.get('armwpp'))
        elif(type == "Round Bar Shape"):
            #D1,L
            editmat.rm_d1 = to_float(request.POST.get('rm-d1'))
            editmat.rm_d1_unit = request.POST.get('rm-d1-unit')
            editmat.rm_l = to_float(request.POST.get('rm-l'))
            editmat.rm_l_unit = request.POST.get('rm-l-unit')
        elif(type == "Tube Shape"):
            #D1,D2,L
            editmat.rm_d1 = to_float(request.POST.get('rm-d1'))
            editmat.rm_d1_unit = request.POST.get('rm-d1-unit')
            editmat.rm_d2 = to_float(request.POST.get('rm-d2'))
            editmat.rm_d2_unit = request.POST.get('rm-d2-unit')
            editmat.rm_l = to_float(request.POST.get('rm-l'))
            editmat.rm_l_unit = request.POST.get('rm-l-unit')
        elif(type == "Flat Bar / Plate / Sheet"):
            #T,W,L
            editmat.rm_t = to_float(request.POST.get('rm-t'))
            editmat.rm_t_unit = request.POST.get('rm-t-unit')
            editmat.rm_w = to_float(request.POST.get('rm-w'))
            editmat.rm_w_unit = request.POST.get('rm-w-unit')
            editmat.rm_l = to_float(request.POST.get('rm-l'))
            editmat.rm_l_unit = request.POST.get('rm-l-unit')
        editmat.supplier = request.POST.get('material-supplier')
        editmat.base_material_price = to_float(request.POST.get('base-material-price'))
        editmat.shipping_cost = to_float(request.POST.get('material-shipping-cost'))
        editmat.fpwp = to_float(request.POST.get('fpwp'))
        editmat.save()

        editmat.armwpp = to_float(request.POST.get('armwpp'))
    return redirect('/part_info/'+sl_no+"?message=1&tab=4")

@login_required(login_url='/login')
def edit_material_remove(request, sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    delmat = Material.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
    clean_material(delmat)
    delmat.save()
    return redirect('/part_info/'+str(sl_no))

@login_required(login_url='/login')
def edit_burden_rate(request, sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    burden_rate = Burden_Rate.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
    part = Part_Header.objects.get(pk=sl_no)
    context = {
            'part': part,
            'burden_rate': burden_rate
    }
    return render(request, 'rfqsite/edit_burden_rate.html', context)

@login_required(login_url='/login')
def edit_burden_rate_confirm(request):
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    sl_no = request.POST.get('sl-no')
    if request.method == 'POST':
        editbd = Burden_Rate.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
        editbd.material = to_float(request.POST.get('rate-material'))
        editbd.hardware = to_float(request.POST.get('rate-hardware'))
        editbd.subcontract = to_float(request.POST.get('rate-subcontract'))
        editbd.sp = to_float(request.POST.get('rate-sp'))
        editbd.save()
    return redirect('/part_info/'+sl_no+"?message=1&tab=6")

def unNan(data):
    if(data == 'NaN' or data == '' or data == None):
        return 0
    else:
        return data

@login_required(login_url='/login')
def data_collect(request):
    if request.method == 'GET':
        sl_no = request.GET.get('sl_no')
        editO = Output.objects.get(sl_no=Part_Header.objects.get(pk=sl_no))
        editO.ccs_ewp = unNan(request.GET.get('ccs_ewp'))
        editO.total_cost = unNan(request.GET.get('total_cost'))
        editO.mtl_cost = unNan(request.GET.get('mtl_cost'))
        editO.spl_process_cost = unNan(request.GET.get('spl_process_cost'))
        editO.total_hardware_cost = unNan(request.GET.get('total_hardware_cost'))
        editO.total_machine_cost = unNan(request.GET.get('total_machine_cost'))
        editO.material_burden_cost = unNan(request.GET.get('material_burden_cost'))
        editO.sp_burden_cost = unNan(request.GET.get('sp_burden_cost'))
        editO.hardware_burden_cost = unNan(request.GET.get('hardware_burden_cost'))
        editO.material_cost = unNan(request.GET.get('material_cost'))
        editO.surface_treatment_cost = unNan(request.GET.get('surface_treatment_cost'))
        editO.hardware_cost = unNan(request.GET.get('hardware_cost'))
        editO.total_manufacturing_cost = unNan(request.GET.get('total_manufacturing_cost'))
        editO.save()
        id = request.GET.getlist('arr_mc_id[]')
        value = request.GET.getlist('arr_mc_mcrftp[]')
        for i in range(len(id)):
            act = ACT_Set.objects.get(id=id[i])
            mc = MC_Set.objects.get(sl_no=sl_no,act_id=act)
            mc.mcrftp = unNan(value[i])
            mc.save()
    return HttpResponse("OK")

class Part_Tree:
    def __init__(self, base_level, current_level):
        self.base_level = base_level
        self.current_level = current_level
        self.sl_no = self.base_level.sl_no
        self.no = self.base_level.no
        self.name = self.base_level.name
        self.type = self.base_level.type
        self.isOpen = False
        self.children = []
    def set_tree(self):
        children = Part_Header.objects.filter(parent_sl_no=self.base_level)
        for child in children:
            pt = Part_Tree(child,self.current_level)
            self.children.append(pt)
        for child in self.children:
            child.set_tree()
    def set_open(self):
        self.isOpen = True
        poc = []
        lcl = self.current_level
        level = self.current_level.level
        while(level > 0):
            poc.append(lcl)
            lcl = lcl.parent_sl_no
            level = lcl.level
        path = self.children
        while(level < self.current_level.level):
            for child in path:
                if child.base_level in poc:
                    child.isOpen = True
                    level = child.base_level.level
                    path = child.children
                    break
    def get_path(self):
        poc = []
        lcl = self.current_level
        level = self.current_level.level
        while(level > 0):
            poc.append(lcl)
            lcl = lcl.parent_sl_no
            level = lcl.level
        poc.append(lcl)
        return poc



#################################################################################
def get_perm(user):
    return user.extenduser.role.permission

@login_required(login_url='/login')
def user_table(request):
    if(get_perm(request.user) != 0):
        return HttpResponseNotFound("Access Denied")
    users = User.objects.all()
    context = {
            'users': users
    }
    return render(request, 'rfqsite/user_table.html', context)

@login_required(login_url='/login')
def add_user(request):
    if(get_perm(request.user) != 0):
        return HttpResponseNotFound("Access Denied")
    roles = Roles.objects.all()
    if request.method == 'POST':
        alluser = User.objects.all()
        for user in alluser:
            if(request.POST.get('username') == user.username):
                message = "username already exist"
                context = {
                'message': message,
                'roles': roles
                }
                return redirect('/user_table/?message=0')
            else:
                newUser = User.objects.create_user(request.POST.get('username'), '', request.POST.get('password'))
                newUser.first_name = request.POST.get('first-name')
                newUser.last_name = request.POST.get('last-name')
                role = Roles.objects.get(permission=request.POST.get('role'))
                ext = ExtendUser(user=newUser,role=role)
                ext.save()
                newUser.save()
                return redirect('/user_table/?message=1')
    else:
        context = {
                'roles': roles
        }
    return render(request, 'rfqsite/add_user.html', context)

@login_required(login_url='/login')
def edit_user(request,username):
    if(get_perm(request.user) != 0):
        return HttpResponseNotFound("Access Denied")
    roles = Roles.objects.all()
    user = User.objects.get(username=username)
    context = {
        'fuser': user,
        'roles': roles
    }
    return render(request, 'rfqsite/edit_user.html', context)

@login_required(login_url='/login')
def edit_user_confirm(request):
    if(get_perm(request.user) != 0):
        return HttpResponseNotFound("Access Denied")
    if request.method == 'POST':
        user = User.objects.get(username=request.POST.get('username'))
        user.first_name = username=request.POST.get('first-name')
        user.last_name = username=request.POST.get('last-name')
        if(request.POST.get('reset-password') != None):
            user.set_password(request.POST.get('password'))
        ext = ExtendUser.objects.get(user=user)
        ext.role = Roles.objects.get(permission=request.POST.get('role'))
        user.save()
        ext.save()
        return redirect('/user_table/?message=1')

@login_required(login_url='/login')
def validate_user(request):
    alluser = User.objects.all()
    for user in alluser:
        if(request.GET.get('username') == user.username):
            data = {'isUsed': True}
            return JsonResponse(data)
    data = {'isUsed': False}
    return JsonResponse(data)

# import requests

# @login_required(login_url='/login')
# def fetch_data(request, tracker_no):
#     rfqp = Part_Header.objects.filter(tracker_no=tracker_no)
#     parts = [x.sl_no for x in rfqp]
#     print(request.get_host())
#     for part in parts:
#         path = str(request.get_host())+"/part_info/"+str(part)
#         response = requests.get(path, headers=None)
#     data = {'data_set': True}
#     return JsonResponse(data)

@login_required(login_url='/login')
def remove_user(request,username):
    if(get_perm(request.user) != 0):
        return HttpResponseNotFound("Access Denied")
    duser = User.objects.get(username=username)
    duser.delete()
    return redirect('/user_table/?message=1')

@login_required(login_url='/login')
def remove_part(request, sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    dpart = Part_Header.objects.get(sl_no=sl_no)
    rfq = dpart.tracker_no
    dpart.delete()
    return redirect('/part_table/'+str(rfq.tracker_no)+'?message=1')

@login_required(login_url='/login')
def master_table(request):
    if(get_perm(request.user) != 0):
        return HttpResponseNotFound("Access Denied")
    sp_table = SP_Master.objects.all()
    mc_table = MC_Master.objects.all()
    context = {
        'sp_masters': sp_table,
        'mc_masters': mc_table
    }
    return render(request, 'rfqsite/master_table.html', context)

@login_required(login_url='/login')
def add_mc_master(request):
    if(get_perm(request.user) != 0):
        return HttpResponseNotFound("Access Denied")
    if request.method == 'POST':
        name = request.POST.get('name')
        for obj in MC_Master.objects.filter(name=name):
            return redirect('/master_table/?message=0')
        sm = MC_Master(name=name)
        sm.save()
        return redirect('/master_table/?message=1')
    context = {
    }
    return render(request, 'rfqsite/add_mc_master.html', context)

@login_required(login_url='/login')
def add_sp_master(request):
    if(get_perm(request.user) != 0):
        return HttpResponseNotFound("Access Denied")
    if request.method == 'POST':
        name = request.POST.get('name')
        for obj in SP_Master.objects.filter(name=name):
            return redirect('/master_table/?message=0')
        sm = SP_Master(name=name)
        sm.save()
        return redirect('/master_table/?message=1')
    context = {
    }
    return render(request, 'rfqsite/add_sp_master.html', context)

@login_required(login_url='/login')
def edit_sp_set(request,sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    part = Part_Header.objects.get(sl_no=sl_no)
    sp_set = SP_Set.objects.filter(sl_no=part).order_by('sp_id')
    context = {
            'part': part,
            'sp_set': sp_set
    }
    return render(request, 'rfqsite/edit_sp_set.html', context)

@login_required(login_url='/login')
def edit_sp_set_confirm(request):
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    if request.method == 'POST':
        sl_no = request.POST.get('sl-no')
        sp = SP_Set.objects.filter(sl_no=sl_no)
        for item in request.POST:
            if('sp-rate' in item):
                id = int(item.replace('sp-rate-',''))
                esp1 = SP_Set.objects.get(sl_no=sl_no,sp_id=SP_Master.objects.get(id=id))
                esp1.rate = request.POST.get(item)
                esp1.save()
            elif('sp-spec' in item):
                id = int(item.replace('sp-spec-',''))
                esp2 = SP_Set.objects.get(sl_no=sl_no,sp_id=SP_Master.objects.get(id=id))
                esp2.spec = request.POST.get(item)
                esp2.save()
        return redirect('/part_info/'+str(sl_no)+'/?message=1&tab=6')

@login_required(login_url='/login')
def select_sp_set(request,sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    part = Part_Header.objects.get(sl_no=sl_no)
    sp_set = SP_Set.objects.filter(sl_no=sl_no)
    sp_master = SP_Master.objects.all()
    select_id = [x.sp_id.id for x in sp_set]
    context = {
            'part': part,
            'sp_set': sp_set,
            'sp_master': sp_master,
            'select_id': select_id
    }
    return render(request, 'rfqsite/select_sp_set.html', context)

@login_required(login_url='/login')
def select_sp_set_confirm(request):
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    if request.method == 'POST':
        sl_no = request.POST.get('sl-no')
        sp_master = SP_Master.objects.all()
        edit = False
        for sp in sp_master:
            id = sp.id
            if(request.POST.get(str(id)) == None):
                delsp = SP_Set.objects.filter(sl_no=sl_no,sp_id=sp)
                delsp.delete()
            elif(request.POST.get(str(id)) == 'on'):
                if(SP_Set.objects.filter(sp_id=sp,sl_no=sl_no)):
                    pass
                else:
                    newsp = SP_Set(sl_no=Part_Header.objects.get(sl_no=sl_no),sp_id=sp)
                    newsp.rate = sp.rate
                    newsp.save()
                    edit = True
        if(edit):
            return redirect('/part_info/'+str(sl_no)+'/?final_page=edit_sp_set&data_set=')
        return redirect('/part_info/'+str(sl_no)+'/?message=1&tab=6')

@login_required(login_url='/login')
def select_act_set(request,tracker_no):
    if(not RFQ.objects.filter(pk=tracker_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    project = RFQ.objects.get(tracker_no=tracker_no)
    act_set = ACT_Set.objects.filter(tracker_no=project)
    mc_master = MC_Master.objects.all()
    select_id = [x.mc_id.id for x in act_set]
    context = {
            'project': project,
            'act_set': act_set,
            'mc_master': mc_master,
            'select_id': select_id
    }
    return render(request, 'rfqsite/select_act_set.html', context)


@login_required(login_url='/login')
def select_act_set_confirm(request):
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    if request.method == 'POST':
        tracker_no = request.POST.get('tracker-no')
        mc_master = MC_Master.objects.all()
        refresh = []
        edit = False
        for mc in mc_master:
            id = mc.id
            if(request.POST.get(str(id)) == None):
                delact = ACT_Set.objects.filter(tracker_no=tracker_no,mc_id=mc)
                act = None
                if(delact):
                    act = delact[0]
                sl_in = MC_Set.objects.filter(act_id=act)
                for mc in sl_in:
                    if mc.sl_no.sl_no not in refresh:
                        refresh.append(mc.sl_no.sl_no)
                delact.delete()
            elif(request.POST.get(str(id)) == 'on'):
                if(ACT_Set.objects.filter(tracker_no=tracker_no,mc_id=mc)):
                    pass
                else:
                    newact = ACT_Set(tracker_no=RFQ.objects.get(tracker_no=tracker_no),mc_id=mc)
                    newact.rate = mc.rate
                    edit = True
                    newact.save()
        if(len(refresh) > 0):
            data_str = ''
            data_set = [str(x) for x in refresh]
            for x in data_set:
                data_str = data_str + x + ","
            data_str = data_str[:len(data_str)-1]
            final_page = 'part_table'
            if(edit):
                final_page = 'edit_act_set'
            return redirect('/part_info/'+str(refresh[0])+'/?final_page='+final_page+'&data_set='+data_str)
        if(edit):
            return redirect('/edit_act_set/'+str(tracker_no)+'/?message=1')
        return redirect('/part_table/'+str(tracker_no)+'/?message=1')

@login_required(login_url='/login')
def edit_act_set(request,tracker_no):
    if(not RFQ.objects.filter(pk=tracker_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    project = RFQ.objects.get(tracker_no=tracker_no)
    act_set = ACT_Set.objects.filter(tracker_no=project).order_by('mc_id')
    context = {
            'project': project,
            'act_set': act_set
    }
    return render(request, 'rfqsite/edit_act_set.html', context)

@login_required(login_url='/login')
def edit_act_set_confirm(request):
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    if request.method == 'POST':
        tracker_no = request.POST.get('tracker-no')
        act = ACT_Set.objects.filter(tracker_no=tracker_no)
        refresh = []
        for item in request.POST:
            if('act-rate' in item):
                id = int(item.replace('act-rate-',''))
                eact = ACT_Set.objects.get(tracker_no=tracker_no,mc_id=MC_Master.objects.get(id=id))
                if(eact.rate != float(request.POST.get(item))):
                    eact.rate = request.POST.get(item)
                    sl_in = MC_Set.objects.filter(act_id=eact)
                    for mc in sl_in:
                        if mc.sl_no.sl_no not in refresh:
                            refresh.append(mc.sl_no.sl_no)
                    eact.save()
        if(len(refresh) > 0):
            data_str = ''
            data_set = [str(x) for x in refresh]
            for x in data_set:
                data_str = data_str + x + ","
            data_str = data_str[:len(data_str)-1]
            return redirect('/part_info/'+str(refresh[0])+'/?final_page=part_table&data_set='+data_str)
        return redirect('/part_table/'+str(tracker_no)+'/?message=1')

@login_required(login_url='/login')
def select_mc_set(request,sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    part = Part_Header.objects.get(sl_no=sl_no)
    mc_set = MC_Set.objects.filter(sl_no=sl_no).order_by('act_id')
    act_set = ACT_Set.objects.filter(tracker_no=part.tracker_no)
    select_id = [x.act_id.id for x in mc_set]
    context = {
            'part': part,
            'mc_set': mc_set,
            'act_set': act_set,
            'select_id': select_id
    }
    return render(request, 'rfqsite/select_mc_set.html', context)

@login_required(login_url='/login')
def select_mc_set_confirm(request):
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    if request.method == 'POST':
        sl_no = request.POST.get('sl-no')
        act_set = ACT_Set.objects.filter(tracker_no=RFQ.objects.get(tracker_no=Part_Header.objects.get(sl_no=sl_no).tracker_no.tracker_no))
        refresh = []
        edit = False
        for act in act_set:
            id = act.id
            if(request.POST.get(str(id)) == None):
                delmc = MC_Set.objects.filter(sl_no=sl_no,act_id=act)
                for item in delmc:
                    if item.sl_no not in refresh:
                        refresh.append(item.sl_no)
                delmc.delete()
            elif(request.POST.get(str(id)) == 'on'):
                if(MC_Set.objects.filter(sl_no=sl_no,act_id=act)):
                    pass
                else:
                    newact = MC_Set(sl_no=Part_Header.objects.get(sl_no=sl_no),act_id=act)
                    newact.save()
                    edit = True
        if(edit):
            return redirect('/part_info/'+str(sl_no)+'/?final_page=edit_mc_set&data_set=')
        return redirect('/part_info/'+str(sl_no)+'/?message=1&tab=6')

@login_required(login_url='/login')
def edit_mc_set(request,sl_no):
    if(not Part_Header.objects.filter(pk=sl_no)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    part = Part_Header.objects.get(sl_no=sl_no)
    mc_set = MC_Set.objects.filter(sl_no=part).order_by('act_id')
    context = {
            'part': part,
            'mc_set': mc_set
    }
    return render(request, 'rfqsite/edit_mc_set.html', context)

@login_required(login_url='/login')
def edit_mc_set_confirm(request):
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    if request.method == 'POST':
        sl_no = request.POST.get('sl-no')
        for item in request.POST:
            if('mc-msut' in item):
                id = int(item.replace('mc-msut-',''))
                emc1 = MC_Set.objects.get(sl_no=sl_no,act_id=ACT_Set.objects.get(id=id))
                emc1.msut = request.POST.get(item)
                emc1.save()
            elif('mc-ctpp' in item):
                id = int(item.replace('mc-ctpp-',''))
                emc2 = MC_Set.objects.get(sl_no=sl_no,act_id=ACT_Set.objects.get(id=id))
                emc2.ctpp = request.POST.get(item)
                emc2.save()
        return redirect('/part_info/'+str(sl_no)+'/?message=1&tab=6')

@login_required(login_url='/login')
def fetch_data(request, tracker_no):
    project = RFQ.objects.get(pk=tracker_no)
    project.save()
    data_set = [str(x.sl_no) for x in Part_Header.objects.filter(tracker_no=tracker_no)]
    data_str = ''
    for x in data_set:
        data_str = data_str + x + ','
    data_str = data_str[:len(data_str)-1]
    return redirect('/part_info/'+str(data_set[0])+'/?final_page=rfq_summary&data_set='+data_str)


@login_required(login_url='/login')
def edit_sp_master(request, id):
    if(not SP_Master.objects.filter(pk=id)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    sp_master = SP_Master.objects.get(pk=id)
    context = {
            'sp': sp_master
    }
    return render(request, 'rfqsite/edit_sp_master.html', context)

@login_required(login_url='/login')
def edit_mc_master(request, id):
    if(not MC_Master.objects.filter(pk=id)):
        return HttpResponseNotFound("Page Not Found")
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    mc_master = MC_Master.objects.get(pk=id)
    context = {
            'mc': mc_master
    }
    return render(request, 'rfqsite/edit_mc_master.html', context)

@login_required(login_url='/login')
def edit_sp_master_confirm(request):
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    if request.method == 'POST':
        id = int(request.POST.get('id'))
        sp_master = SP_Master.objects.get(pk=id)
        sp_master.name = request.POST.get('name')
        sp_master.rate = to_float(request.POST.get('rate'))
        sp_master.save()
        return redirect('/master_table/'+'?message=1')

@login_required(login_url='/login')
def edit_mc_master_confirm(request):
    if(get_perm(request.user) > 1):
        return HttpResponseNotFound("Access Denied")
    if request.method == 'POST':
        id = int(request.POST.get('id'))
        mc_master = MC_Master.objects.get(pk=id)
        mc_master.name = request.POST.get('name')
        mc_master.rate = to_float(request.POST.get('rate'))
        mc_master.save()
        return redirect('/master_table/'+'?message=1')

from openpyxl import Workbook
from openpyxl.styles import PatternFill,colors, Alignment, Font, Color
#pip install openpyxl

class Customer_Sheet:

    def __init__(self):
        self.sl_no = 0
        self.no = ''
        self.name = ''
        self.program = ''
        self.forecast_year1 = 0
        self.forecast_year2 = 0
        self.forecast_year3 = 0
        self.forecast_year4 = 0
        self.forecast_year5 = 0
        self.description = ''
        self.material_cost = 0
        self.surface_treatment_cost = 0
        self.hardware_cost = 0
        self.total_manufacturing_cost = 0
        self.ccs_ewp = 0
        self.ebq_customer_qty = 0
        self.aeqfc = 0
        self.ottc = 0
        self.ccs_quote_assumptions = ''
        self.dltiw_fai = ''
        self.dltiw_serial_production = ''
        self.dltiw_production = ''



@login_required(login_url='/login')
def generate_table(request, tracker_no):
    rfq = RFQ.objects.get(pk=tracker_no)
    rfq.last_generate = datetime.now()
    tracker = rfq.ccs_rfq.ccs_tracker_no
    current_year = rfq.current_year
    main_part = Part_Header.objects.filter(tracker_no=rfq,level=0).order_by('sl_no')
    sl_no = 1
    to_excel = []
    for part in main_part:
        sheet = Customer_Sheet()
        sheet.sl_no = sl_no
        sheet.no = part.no
        sheet.name = part.name
        sheet.program = part.program
        forecast = Forecast.objects.get(sl_no=part)
        sheet.forecast_year1 = forecast.forecast_year1
        sheet.forecast_year2 = forecast.forecast_year2
        sheet.forecast_year3 = forecast.forecast_year3
        sheet.forecast_year4 = forecast.forecast_year4
        sheet.forecast_year5 = forecast.forecast_year5
        sheet.aeqfc = forecast.aeqfc
        part_cost = Part_Costing.objects.get(sl_no=part)
        sheet.ccs_quote_assumptions = part_cost.ccs_quote_assumptions
        sheet.ebq_customer_qty = part_cost.ebq_customer_qty
        sheet.dltiw_fai = part_cost.dltiw_fai
        sheet.dltiw_production = part_cost.dltiw_production
        sheet.dltiw_serial_production = part_cost.dltiw_serial_production
        all_part = []
        c = [part]
        while(len(c) > 0):
            for item in c:
                all_part.append(item)
                fo = Part_Header.objects.filter(parent_sl_no=item)
                for item2 in fo:
                    c.append(item2)
                c.remove(item)
        for single_part in all_part:
            output = Output.objects.get(sl_no=single_part)
            part_costing = Part_Costing.objects.get(sl_no=single_part)
            material = Material.objects.get(sl_no=single_part)
            if(material.description != ''):
                sheet.description = sheet.description + material.description + ', '
            sheet.ottc += part_costing.ottc
            sheet.material_cost += output.material_cost
            sheet.surface_treatment_cost += output.surface_treatment_cost
            sheet.hardware_cost += output.hardware_cost
            sheet.total_manufacturing_cost += output.total_manufacturing_cost
            sheet.ccs_ewp += output.ccs_ewp
        if(len(sheet.description) > 2):
            sheet.description = sheet.description[:len(sheet.description)-2]
        sl_no += 1
        to_excel.append(sheet)
    ### excel ###
    header = ['SL NO.', 'Part No.', 'Part Name', 'Program', 'Forecast '+str(current_year), 'Forecast '+str(current_year+1), 'Forecast '+str(current_year+2), 'Forecast '+str(current_year+3), 'Forecast '+str(current_year+4), 'Material Descrition', 'Matrial Cost DDP CCS', 'Total Hardwares cost', 'Surface treatment cost', 'Total manufacturing cost', 'CCS price per Part/assy Ex_works BKK Thailand', 'EBQ QTY (Prodnlot)', 'Minimum firm order quantity/lot', 'One time tooling cost', 'CCS Quote Assumptions', 'Delivery Lead Time in weeks FAI', 'Delivery Lead Time in weeks Serial Production', 'Delivery Lead Time in weeks Production']
    wb = Workbook()
    ws = wb.active
    ws.title = tracker
    for i in range(len(header)):
        ws.cell(row=1, column=i+1).value = header[i]
        ws.cell(row=1, column=i+1).alignment = Alignment(wrap_text=True,vertical='top')
        ws.cell(row=1, column=i+1).font = Font(color="000000", italic=True)
        ws.cell(row=1, column=i+1).fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
    for i in range(len(to_excel)):
        ws.cell(row=i+2,column=1).value = to_excel[i].sl_no
        ws.cell(row=i+2,column=2).value = to_excel[i].no
        ws.cell(row=i+2,column=3).value = to_excel[i].name
        ws.cell(row=i+2,column=4).value = to_excel[i].program
        ws.cell(row=i+2,column=5).value = to_excel[i].forecast_year1
        ws.cell(row=i+2,column=6).value = to_excel[i].forecast_year2
        ws.cell(row=i+2,column=7).value = to_excel[i].forecast_year3
        ws.cell(row=i+2,column=8).value = to_excel[i].forecast_year4
        ws.cell(row=i+2,column=9).value = to_excel[i].forecast_year5
        ws.cell(row=i+2,column=10).value = to_excel[i].description
        #### Dollar
        ws.cell(row=i+2,column=11).value = to_excel[i].material_cost
        ws.cell(row=i+2,column=11).number_format = '$ 0.00'
        ws.cell(row=i+2,column=12).value = to_excel[i].hardware_cost
        ws.cell(row=i+2,column=12).number_format = '$ 0.00'
        ws.cell(row=i+2,column=13).value = to_excel[i].surface_treatment_cost
        ws.cell(row=i+2,column=13).number_format = '$ 0.00'
        ws.cell(row=i+2,column=14).value = to_excel[i].total_manufacturing_cost
        ws.cell(row=i+2,column=14).number_format = '$ 0.00'
        ws.cell(row=i+2,column=15).value = to_excel[i].ccs_ewp
        ws.cell(row=i+2,column=15).number_format = '$ 0.00'
        ws.cell(row=i+2,column=16).value = to_excel[i].ebq_customer_qty
        ws.cell(row=i+2,column=16).number_format = '$ 0.00'
        ws.cell(row=i+2,column=17).value = to_excel[i].aeqfc
        ws.cell(row=i+2,column=17).number_format = '$ 0.00'
        ws.cell(row=i+2,column=18).value = to_excel[i].ottc
        ws.cell(row=i+2,column=18).number_format = '$ 0.00'
        ######
        ws.cell(row=i+2,column=19).value = to_excel[i].ccs_quote_assumptions
        ws.cell(row=i+2,column=20).value = to_excel[i].dltiw_fai
        ws.cell(row=i+2,column=21).value = to_excel[i].dltiw_serial_production
        ws.cell(row=i+2,column=22).value = to_excel[i].dltiw_production
    wb.save('media/gen/'+tracker+'.xlsx')
    rfq.customer_file_path = 'gen/'+tracker+'.xlsx'
    rfq.save()
    return redirect('/rfq_summary/'+str(tracker_no)+'/?message=3')
